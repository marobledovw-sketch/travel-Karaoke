#!/usr/bin/env python3
"""
Rellena la columna Imagen_URL de catalogo_karaoke.xlsx buscando la carátula
de cada canción en la API pública de iTunes, UNA SOLA VEZ por adelantado.

Por qué existe este script:
La app busca carátulas en directo (en el móvil de cada usuario) contra la
API de iTunes, que tiene un límite de peticiones no oficial bastante bajo.
En el evento real, todos los móviles comparten la IP pública del wifi del
local, así que si mucha gente navega el catálogo a la vez, iTunes puede
empezar a bloquear esa IP compartida y dejar de servir carátulas a todos.

Ejecutando este script UNA VEZ (antes del evento, desde tu propio Mac) se
resuelven las carátulas de antemano y se guardan en el Excel. Así, durante
el evento, la app apenas necesita llamar a iTunes en directo: la mayoría de
canciones ya llevan su Imagen_URL puesta.

Uso:
    python3 -m pip install --quiet openpyxl
    python3 enrich_covers.py

Se puede parar (Ctrl+C) y volver a lanzar más tarde: guarda el progreso
cada 20 canciones y solo busca las que todavía no tengan Imagen_URL.
"""
import json
import time
import urllib.request
import urllib.parse
import urllib.error

from openpyxl import load_workbook

ARCHIVO = "catalogo_karaoke.xlsx"
HOJA = "Catalogo"
PAUSA_SEGUNDOS = 2.5          # espera entre peticiones, para no saturar iTunes
GUARDAR_CADA = 20             # filas procesadas antes de guardar progreso
REINTENTOS_MAX = 3


def buscar_caratula(termino):
    url = "https://itunes.apple.com/search?" + urllib.parse.urlencode({
        "term": termino, "media": "music", "limit": 1
    })
    for intento in range(1, REINTENTOS_MAX + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as res:
                data = json.loads(res.read().decode("utf-8"))
                resultados = data.get("results") or []
                if resultados and resultados[0].get("artworkUrl100"):
                    return resultados[0]["artworkUrl100"].replace("100x100", "300x300")
                return None
        except urllib.error.HTTPError as e:
            if e.code in (403, 429) and intento < REINTENTOS_MAX:
                espera = 8 * intento
                print(f"    iTunes devolvió {e.code}, esperando {espera}s y reintentando...")
                time.sleep(espera)
                continue
            print(f"    iTunes devolvió {e.code}, se deja sin carátula.")
            return None
        except Exception as e:
            print(f"    Error de red ({e}), se deja sin carátula.")
            return None
    return None


def main():
    wb = load_workbook(ARCHIVO)
    ws = wb[HOJA]
    headers = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}

    total = ws.max_row - 1
    encontradas = 0
    ya_tenian = 0
    sin_resultado = 0

    print(f"Procesando {total} canciones de '{ARCHIVO}'...\n")

    for fila in range(2, ws.max_row + 1):
        titulo = ws.cell(row=fila, column=col["Titulo"]).value or ""
        artista = ws.cell(row=fila, column=col["Artista"]).value or ""
        imagen_actual = ws.cell(row=fila, column=col["Imagen_URL"]).value

        if imagen_actual:
            ya_tenian += 1
            continue

        termino = f"{artista} {titulo}".strip() or titulo
        print(f"[{fila - 1}/{total}] {termino}")
        url_caratula = buscar_caratula(termino)

        if url_caratula:
            ws.cell(row=fila, column=col["Imagen_URL"], value=url_caratula)
            encontradas += 1
        else:
            sin_resultado += 1

        if (fila - 1) % GUARDAR_CADA == 0:
            wb.save(ARCHIVO)
            print(f"  -- progreso guardado ({fila - 1}/{total}) --")

        time.sleep(PAUSA_SEGUNDOS)

    wb.save(ARCHIVO)
    print("\nListo.")
    print(f"  Ya tenían carátula puesta a mano: {ya_tenian}")
    print(f"  Carátulas nuevas encontradas:     {encontradas}")
    print(f"  Sin carátula encontrada:          {sin_resultado}")
    print(f"\nGuardado en {ARCHIVO}. Vuelve a pedirme que regenere")
    print("catalogo_data.js a partir de este Excel para que la app las use.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrumpido. Puedes volver a ejecutar el script más tarde:")
        print("solo buscará las canciones que todavía no tengan Imagen_URL.")
