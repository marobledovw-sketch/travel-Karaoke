#!/usr/bin/env python3
"""
Descarga a la carpeta local `caratulas/` la imagen de cada canción que ya
tenga una Imagen_URL (resuelta antes por enrich_covers.py), y anota en el
Excel qué canciones ya están sincronizadas con esa carpeta (columna
Imagen_Local), para no volver a descargar la misma imagen dos veces.

Por qué existe este script:
Hasta ahora la app carga cada carátula en directo desde el CDN de Apple
(mzstatic.com), lo que depende de que esa red esté disponible y puede ir
más lento. Teniendo las imágenes ya descargadas en la propia carpeta del
proyecto (que se sube a GitHub Pages junto con index.html/admin.html), la
app las sirve como archivos propios: más rápido y sin depender de iTunes
en directo durante el evento.

Uso:
    python3 -m pip install --quiet openpyxl
    python3 descargar_caratulas.py

Se puede parar (Ctrl+C) y volver a lanzar más tarde: solo descarga las
canciones cuya Imagen_Local todavía no apunte a un archivo que ya existe.
"""
import os
import time
import urllib.request
import urllib.error

from openpyxl import load_workbook

ARCHIVO = "catalogo_karaoke.xlsx"
HOJA = "Catalogo"
CARPETA = "caratulas"
PAUSA_SEGUNDOS = 0.3
GUARDAR_CADA = 30
REINTENTOS_MAX = 2


def descargar(url, ruta_local):
    for intento in range(1, REINTENTOS_MAX + 1):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as res:
                datos = res.read()
            with open(ruta_local, "wb") as f:
                f.write(datos)
            return True
        except Exception as e:
            if intento < REINTENTOS_MAX:
                time.sleep(3)
                continue
            print(f"    Error descargando: {e}")
            return False


def main():
    os.makedirs(CARPETA, exist_ok=True)
    wb = load_workbook(ARCHIVO)
    ws = wb[HOJA]
    headers = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers)}

    if "Imagen_Local" not in col:
        nueva_col = len(headers) + 1
        ws.cell(row=1, column=nueva_col, value="Imagen_Local")
        col["Imagen_Local"] = nueva_col

    total = ws.max_row - 1
    descargadas = 0
    ya_sincronizadas = 0
    sin_url = 0
    fallidas = 0

    print(f"Sincronizando carátulas de {total} canciones en '{CARPETA}/'...\n")

    for fila in range(2, ws.max_row + 1):
        id_cancion = ws.cell(row=fila, column=col["ID"]).value
        url = ws.cell(row=fila, column=col["Imagen_URL"]).value
        local_actual = ws.cell(row=fila, column=col["Imagen_Local"]).value

        if not url:
            sin_url += 1
            continue

        nombre_archivo = f"{id_cancion}.jpg"
        ruta_local = os.path.join(CARPETA, nombre_archivo)
        ruta_relativa = f"{CARPETA}/{nombre_archivo}"

        if local_actual == ruta_relativa and os.path.exists(ruta_local) and os.path.getsize(ruta_local) > 0:
            ya_sincronizadas += 1
            continue

        print(f"[{fila - 1}/{total}] {nombre_archivo}")
        if descargar(url, ruta_local):
            ws.cell(row=fila, column=col["Imagen_Local"], value=ruta_relativa)
            descargadas += 1
        else:
            fallidas += 1

        if (fila - 1) % GUARDAR_CADA == 0:
            wb.save(ARCHIVO)
            print(f"  -- progreso guardado ({fila - 1}/{total}) --")

        time.sleep(PAUSA_SEGUNDOS)

    wb.save(ARCHIVO)
    print("\nListo.")
    print(f"  Ya sincronizadas de antes: {ya_sincronizadas}")
    print(f"  Descargadas ahora:         {descargadas}")
    print(f"  Sin carátula (Imagen_URL vacía): {sin_url}")
    print(f"  Fallidas:                  {fallidas}")
    print(f"\nGuardado en {ARCHIVO}. Vuelve a pedirme que regenere")
    print("catalogo_data.js y sincronice el Google Sheet con la columna Imagen_Local.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrumpido. Puedes volver a ejecutar el script más tarde:")
        print("solo descargará las que todavía no tenga en la carpeta.")
